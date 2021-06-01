#Define Library
from tkinter import *
from tkinter import ttk
from tkinter.ttk import Combobox
from PIL import ImageTk, Image
import tkinter as tk
from time import sleep
import sqlite3
import tkinter.messagebox as tkMessageBox
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import sklearn
from sklearn.model_selection import train_test_split
import numpy as np
import pandas as pd
import re
import sklearn
from sklearn.model_selection import cross_val_score
from sklearn.model_selection import train_test_split
from sklearn.ensemble import VotingClassifier
from sklearn import ensemble
from sklearn.ensemble import RandomForestClassifier
from sklearn.neighbors import KNeighborsClassifier
from sklearn.neural_network import MLPClassifier
from sklearn.ensemble import AdaBoostClassifier
from sklearn import tree
from sklearn.naive_bayes import GaussianNB
from sklearn.ensemble import GradientBoostingClassifier
from sklearn import svm
from sklearn.ensemble import ExtraTreesClassifier    
from sklearn import model_selection
from sklearn.ensemble import VotingClassifier
from sklearn.model_selection import cross_val_score
from sklearn import model_selection

#Define Diabetes Data
diabetes=pd.read_csv('diabetes.csv')
x1 = diabetes[['Pregnancies','Glucose','BloodPressure','SkinThickness','Insulin','BMI','DiabetesPedigreeFunction','Age']]
y1 = diabetes['Outcome']
X_train,X_test,y_train,y_test=train_test_split(x1,y1,test_size=30,random_state=0)

from sklearn.model_selection import train_test_split
from mlxtend.classifier import StackingClassifier
from sklearn.ensemble import RandomForestClassifier, GradientBoostingClassifier
from sklearn.linear_model import LogisticRegression, RidgeClassifier
from sklearn import datasets, metrics
import stacked_generalization
from mlxtend.classifier import StackingClassifier
from stacked_generalization import StackedGeneralizer

bclf = LogisticRegression(random_state=1)
# Stage 0 models
clfs = [RandomForestClassifier(n_estimators=40, criterion = 'gini', random_state=1),GradientBoostingClassifier(n_estimators=25, random_state=1),RidgeClassifier(random_state=1)]
sl = StackingClassifier(classifiers=clfs,meta_classifier=bclf)
sl.fit(x1,y1)


#Define Algorithm
from sklearn.ensemble import RandomForestClassifier
rfc=RandomForestClassifier(n_estimators=12)
rfc.fit(x1,y1)


from sklearn.neural_network import MLPClassifier
MLP=MLPClassifier(alpha=0.01,hidden_layer_sizes=(100,),validation_fraction=0.9,shuffle=False)
MLP.fit(X_train,y_train)

from sklearn.neighbors import KNeighborsClassifier
knn=KNeighborsClassifier(n_neighbors=2,p=1,leaf_size=738)
knn.fit(X_train,y_train)
        
from sklearn.ensemble import AdaBoostClassifier
ABC=AdaBoostClassifier(n_estimators=67)
ABC.fit(x1,y1)

from sklearn.tree import DecisionTreeClassifier
DTree=tree.DecisionTreeClassifier(criterion='gini',random_state=60)
DTree.fit(x1,y1)

from sklearn.naive_bayes import GaussianNB
GNB=GaussianNB()
GNB.fit(X_train,y_train)

from sklearn.ensemble import GradientBoostingClassifier
gbc=GradientBoostingClassifier(n_estimators=4,max_depth=11,random_state=2,min_samples_leaf=2)
gbc.fit(X_train,y_train)

from sklearn import svm
SVM=svm.SVC(kernel='rbf', probability=True)
SVM.fit(X_train,y_train)

from sklearn.ensemble import ExtraTreesClassifier
ETC=ExtraTreesClassifier(random_state = 0, n_jobs = -1, n_estimators = 100, max_depth = 3)
ETC.fit(X_train,y_train)
import statistics
from openpyxl import *
wb = load_workbook('Vital_Signs.xlsx')
wb.save('Vital_Signs.xlsx')
sheet = wb.active
def Insert():
    #wb = load_workbook('practice.xlsx')
    #wb.save('excel.xlsx')


    def excel():
        
            # resize the width of columns in 
            # excel spreadsheet 
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 10
        sheet.column_dimensions['C'].width = 10
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 20
        sheet.column_dimensions['F'].width = 40
        sheet.column_dimensions['G'].width = 50

            # write given data to an excel spreadsheet 
            # at particular location 
        sheet.cell(row=1, column=1).value = "Name"
        sheet.cell(row=1, column=2).value = "Age"
        sheet.cell(row=1, column=3).value = "Sex"
        sheet.cell(row=1, column=4).value = "Tell Phone"
        sheet.cell(row=1, column=5).value = "Address"
        sheet.cell(row=1, column=6).value = "Cause of referral"
        sheet.cell(row=1, column=7).value = "Admission Date"


    # Function to set focus (cursor) 
    def focus1(event):
        
            # set focus on the course_field box 
        course_field.focus_set() 


    # Function to set focus 
    def focus2(event): 
            # set focus on the sem_field box 
        sem_field.focus_set() 


    # Function to set focus 
    def focus3(event): 
            # set focus on the form_no_field box 
        form_no_field.focus_set() 


    # Function to set focus 
    def focus4(event): 
            # set focus on the contact_no_field box 
        contact_no_field.focus_set() 


    # Function to set focus 
    def focus5(event): 
            # set focus on the email_id_field box 
        email_id_field.focus_set() 


    # Function to set focus 
    def focus6(event): 
            # set focus on the address_field box 
        address_field.focus_set() 


    # Function for clearing the 
    # contents of text entry boxes 
    def clear():
        
            
            # clear the content of text entry box 
        name_field.delete(0, END) 
        course_field.delete(0, END) 
        sem_field.delete(0, END) 
        form_no_field.delete(0, END) 
        contact_no_field.delete(0, END) 
        email_id_field.delete(0, END) 
        address_field.delete(0, END) 


    # Function to take data from GUI 
    # window and write to an excel file 
    def insert():
        
            
            # if user not fill any entry 
            # then print "empty input" 
        if (name_field.get() == "" and
                course_field.get() == "" and
                sem_field.get() == "" and
                form_no_field.get() == "" and
                contact_no_field.get() == "" and
                email_id_field.get() == "" and
                address_field.get() == ""): 
                            
                print("empty input") 

        else:
                

                    # assigning the max row and max column 
                    # value upto which data is written 
                    # in an excel sheet to the variable 
                current_row = sheet.max_row 
                current_column = sheet.max_column 

                    # get method returns current text 
                    # as string which we write into 
                    # excel spreadsheet at particular location 
                sheet.cell(row=current_row + 1, column=1).value = name_field.get() 
                sheet.cell(row=current_row + 1, column=2).value = course_field.get() 
                sheet.cell(row=current_row + 1, column=3).value = sem_field.get() 
                sheet.cell(row=current_row + 1, column=4).value = form_no_field.get() 
                sheet.cell(row=current_row + 1, column=5).value = contact_no_field.get() 
                sheet.cell(row=current_row + 1, column=6).value = email_id_field.get() 
                sheet.cell(row=current_row + 1, column=7).value = address_field.get() 

                    # save the file 
                wb.save('Vital_Signs.xlsx')

                    # set focus on the name_field box 
                name_field.focus_set() 

                    # call the clear() function 
                clear() 


    # Driver code 
    if __name__ == "__main__":
        
            # create a GUI window
        root = Tk() 

            # set the background colour of GUI window 
        root.configure(background='light green') 

            # set the title of GUI window 
        root.title("registration form") 

            # set the configuration of GUI window 
        root.geometry("500x300") 

        excel() 

            # create a Form label 
        heading = Label(root, text="Form", bg="light green") 

            # create a Name label 
        name = Label(root, text="نام و نام خانوادگي", bg="light green") 

            # create a Course label 
        course = Label(root, text="سن", bg="light green") 

            # create a Semester label 
        sem = Label(root, text="جنسيت", bg="light green") 

            # create a Form No. lable 
        form_no = Label(root, text="شماره تلفن", bg="light green") 

            # create a Contact No. label 
        contact_no = Label(root, text="آدرس", bg="light green") 

            # create a Email id label 
        email_id = Label(root, text="علت مراجعه", bg="light green") 

            # create a address label 
        address = Label(root, text="تاريخ پذيرش", bg="light green") 

            # grid method is used for placing 
            # the widgets at respective positions 
            # in table like structure . 
        heading.grid(row=0, column=1) 
        name.grid(row=1, column=0) 
        course.grid(row=2, column=0) 
        sem.grid(row=3, column=0) 
        form_no.grid(row=4, column=0) 
        contact_no.grid(row=5, column=0) 
        email_id.grid(row=6, column=0) 
        address.grid(row=7, column=0) 

            # create a text entry box 
            # for typing the information 
        name_field = Entry(root) 
        course_field = Entry(root) 
        sem_field = Entry(root) 
        form_no_field = Entry(root) 
        contact_no_field = Entry(root) 
        email_id_field = Entry(root) 
        address_field = Entry(root) 

            # bind method of widget is used for 
            # the binding the function with the events 

            # whenever the enter key is pressed 
            # then call the focus1 function 
        name_field.bind("<Return>", focus1) 

            # whenever the enter key is pressed 
            # then call the focus2 function 
        course_field.bind("<Return>", focus2) 

            # whenever the enter key is pressed 
            # then call the focus3 function 
        sem_field.bind("<Return>", focus3) 

            # whenever the enter key is pressed 
            # then call the focus4 function 
        form_no_field.bind("<Return>", focus4) 

            # whenever the enter key is pressed 
            # then call the focus5 function 
        contact_no_field.bind("<Return>", focus5) 

            # whenever the enter key is pressed 
            # then call the focus6 function 
        email_id_field.bind("<Return>", focus6) 

            # grid method is used for placing 
            # the widgets at respective positions 
            # in table like structure . 
        name_field.grid(row=1, column=1, ipadx="100") 
        course_field.grid(row=2, column=1, ipadx="100") 
        sem_field.grid(row=3, column=1, ipadx="100") 
        form_no_field.grid(row=4, column=1, ipadx="100") 
        contact_no_field.grid(row=5, column=1, ipadx="100") 
        email_id_field.grid(row=6, column=1, ipadx="100") 
        address_field.grid(row=7, column=1, ipadx="100") 

            # call excel function
        excel()
            # create a Submit Button and place into the root window 
        submit = Button(root, text="ثبت", fg="Black",bg="Red", command=insert)                                                     
        submit.grid(row=8, column=1)
        root.mainloop()

def new_user():
    
    root = Tk()
    root.title("ثبت کاربر جديد")
    width = 500
    height = 400
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    x = (screen_width/2) - (width/2)
    y = (screen_height/2) - (height/2)
    root.geometry("%dx%d+%d+%d" % (width, height, x, y))
    root.resizable(0, 0)
    root.config(bg="#6666ff")

    #============================VARIABLES===================================
    FIRSTNAME = StringVar()
    LASTNAME = StringVar()
    GENDER = StringVar()
    AGE = StringVar()
    ADDRESS = StringVar()
    CONTACT = StringVar()



    #============================METHODS=====================================

    def Database():
        
        conn = sqlite3.connect("pythontut.db")
        cursor = conn.cursor()
        cursor.execute("CREATE TABLE IF NOT EXISTS `member` (mem_id INTEGER NOT NULL  PRIMARY KEY AUTOINCREMENT, firstname TEXT, lastname TEXT, gender TEXT, age TEXT, address TEXT, contact TEXT)")
        cursor.execute("SELECT * FROM `member` ORDER BY `lastname` ASC")
        fetch = cursor.fetchall()
        for data in fetch:
            tree.insert('', 'end', values=(data))
        cursor.close()
        conn.close()

    def SubmitData():
        
        if  FIRSTNAME.get() == "" or LASTNAME.get() == "" or GENDER.get() == "" or AGE.get() == "" or ADDRESS.get() == "" or CONTACT.get() == "":
            result = tkMessageBox.showwarning('', 'Please Complete The Required Field', icon="warning")
        else:
            tree.delete(*tree.get_children())
            conn = sqlite3.connect("pythontut.db")
            cursor = conn.cursor()
            cursor.execute("INSERT INTO `member` (firstname, lastname, gender, age, address, contact) VALUES(?, ?, ?, ?, ?, ?)", (str(FIRSTNAME.get()), str(LASTNAME.get()), str(GENDER.get()), int(AGE.get()), str(ADDRESS.get()), str(CONTACT.get())))
            conn.commit()
            cursor.execute("SELECT * FROM `member` ORDER BY `lastname` ASC")
            fetch = cursor.fetchall()
            for data in fetch:
                tree.insert('', 'end', values=(data))
            cursor.close()
            conn.close()
            FIRSTNAME.set("")
            LASTNAME.set("")
            GENDER.set("")
            AGE.set("")
            ADDRESS.set("")
            CONTACT.set("")

    def UpdateData():
        if GENDER.get() == "":
           result = tkMessageBox.showwarning('', 'Please Complete The Required Field', icon="warning")
        else:
            tree.delete(*tree.get_children())
            conn = sqlite3.connect("pythontut.db")
            cursor = conn.cursor()
            cursor.execute("UPDATE `member` SET `firstname` = ?, `lastname` = ?, `gender` =?, `age` = ?,  `address` = ?, `contact` = ? WHERE `mem_id` = ?", (str(FIRSTNAME.get()), str(LASTNAME.get()), str(GENDER.get()), str(AGE.get()), str(ADDRESS.get()), str(CONTACT.get()), int(mem_id)))
            conn.commit()
            cursor.execute("SELECT * FROM `member` ORDER BY `lastname` ASC")
            fetch = cursor.fetchall()
            for data in fetch:
                tree.insert('', 'end', values=(data))
            cursor.close()
            conn.close()
            FIRSTNAME.set("")
            LASTNAME.set("")
            GENDER.set("")
            AGE.set("")
            ADDRESS.set("")
            CONTACT.set("")
            
        
    def OnSelected(event):
        
        global mem_id, UpdateWindow
        curItem = tree.focus()
        contents =(tree.item(curItem))
        selecteditem = contents['values']
        mem_id = selecteditem[0]
        FIRSTNAME.set("")
        LASTNAME.set("")
        GENDER.set("")
        AGE.set("")
        ADDRESS.set("")
        CONTACT.set("")
        FIRSTNAME.set(selecteditem[1])
        LASTNAME.set(selecteditem[2])
        AGE.set(selecteditem[4])
        ADDRESS.set(selecteditem[5])
        CONTACT.set(selecteditem[6])
        UpdateWindow = Toplevel()
        UpdateWindow.title("ويرايش")
        width = 400
        height = 300
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()
        x = ((screen_width/2) + 450) - (width/2)
        y = ((screen_height/2) + 20) - (height/2)
        UpdateWindow.resizable(0, 0)
        UpdateWindow.geometry("%dx%d+%d+%d" % (width, height, x, y))
        if 'NewWindow' in globals():
            NewWindow.destroy()

        #===================FRAMES==============================
        FormTitle = Frame(UpdateWindow)
        FormTitle.pack(side=TOP)
        ContactForm = Frame(UpdateWindow)
        ContactForm.pack(side=TOP, pady=10)
        RadioGroup = Frame(ContactForm)
        Male = Radiobutton(RadioGroup, text="Male", variable=GENDER, value="Male",  font=('arial', 14)).pack(side=LEFT)
        Female = Radiobutton(RadioGroup, text="Female", variable=GENDER, value="Female",  font=('arial', 14)).pack(side=LEFT)
        
        #===================LABELS==============================
        lbl_title = Label(FormTitle, text="Updating Contacts", font=('arial', 16), bg="orange",  width = 300)
        lbl_title.pack(fill=X)
        lbl_firstname = Label(ContactForm, text="Firstname", font=('arial', 14), bd=5)
        lbl_firstname.grid(row=0, sticky=W)
        lbl_lastname = Label(ContactForm, text="Lastname", font=('arial', 14), bd=5)
        lbl_lastname.grid(row=1, sticky=W)
        lbl_gender = Label(ContactForm, text="Gender", font=('arial', 14), bd=5)
        lbl_gender.grid(row=2, sticky=W)
        lbl_age = Label(ContactForm, text="Age", font=('arial', 14), bd=5)
        lbl_age.grid(row=3, sticky=W)
        lbl_address = Label(ContactForm, text="Address", font=('arial', 14), bd=5)
        lbl_address.grid(row=4, sticky=W)
        lbl_contact = Label(ContactForm, text="Contact", font=('arial', 14), bd=5)
        lbl_contact.grid(row=5, sticky=W)

        #===================ENTRY===============================
        firstname = Entry(ContactForm, textvariable=FIRSTNAME, font=('arial', 14))
        firstname.grid(row=0, column=1)
        lastname = Entry(ContactForm, textvariable=LASTNAME, font=('arial', 14))
        lastname.grid(row=1, column=1)
        RadioGroup.grid(row=2, column=1)
        age = Entry(ContactForm, textvariable=AGE,  font=('arial', 14))
        age.grid(row=3, column=1)
        address = Entry(ContactForm, textvariable=ADDRESS,  font=('arial', 14))
        address.grid(row=4, column=1)
        contact = Entry(ContactForm, textvariable=CONTACT,  font=('arial', 14))
        contact.grid(row=5, column=1)
        

        #==================BUTTONS==============================
        btn_updatecon = Button(ContactForm, text="Update", width=50, command=UpdateData)
        btn_updatecon.grid(row=6, columnspan=2, pady=10)


        
    def DeleteData():
        if not tree.selection():
           result = tkMessageBox.showwarning('', 'Please Select Something First!', icon="warning")
        else:
            result = tkMessageBox.askquestion('', 'Are you sure you want to delete this record?', icon="warning")
            if result == 'yes':
                curItem = tree.focus()
                contents =(tree.item(curItem))
                selecteditem = contents['values']
                tree.delete(curItem)
                conn = sqlite3.connect("pythontut.db")
                cursor = conn.cursor()
                cursor.execute("DELETE FROM `member` WHERE `mem_id` = %d" % selecteditem[0])
                conn.commit()
                cursor.close()
                conn.close()
        
    def AddNewWindow():
        global NewWindow
        FIRSTNAME.set("")
        LASTNAME.set("")
        GENDER.set("")
        AGE.set("")
        ADDRESS.set("")
        CONTACT.set("")
        NewWindow = Toplevel()
        NewWindow.title("Sourcecodester")
        width = 400
        height = 300
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()
        x = ((screen_width/2) - 455) - (width/2)
        y = ((screen_height/2) + 20) - (height/2)
        NewWindow.resizable(0, 0)
        NewWindow.geometry("%dx%d+%d+%d" % (width, height, x, y))
        if 'UpdateWindow' in globals():
            UpdateWindow.destroy()
        
        #===================FRAMES==============================
        FormTitle = Frame(NewWindow)
        FormTitle.pack(side=TOP)
        ContactForm = Frame(NewWindow)
        ContactForm.pack(side=TOP, pady=10)
        RadioGroup = Frame(ContactForm)
        Male = Radiobutton(RadioGroup, text="Male", variable=GENDER, value="Male",  font=('arial', 14)).pack(side=LEFT)
        Female = Radiobutton(RadioGroup, text="Female", variable=GENDER, value="Female",  font=('arial', 14)).pack(side=LEFT)
        
        #===================LABELS==============================
        lbl_title = Label(FormTitle, text="Adding New Contacts", font=('arial', 16), bg="#66ff66",  width = 300)
        lbl_title.pack(fill=X)
        lbl_firstname = Label(ContactForm, text="Firstname", font=('arial', 14), bd=5)
        lbl_firstname.grid(row=0, sticky=W)
        lbl_lastname = Label(ContactForm, text="Lastname", font=('arial', 14), bd=5)
        lbl_lastname.grid(row=1, sticky=W)
        lbl_gender = Label(ContactForm, text="Gender", font=('arial', 14), bd=5)
        lbl_gender.grid(row=2, sticky=W)
        lbl_age = Label(ContactForm, text="Age", font=('arial', 14), bd=5)
        lbl_age.grid(row=3, sticky=W)
        lbl_address = Label(ContactForm, text="Address", font=('arial', 14), bd=5)
        lbl_address.grid(row=4, sticky=W)
        lbl_contact = Label(ContactForm, text="Contact", font=('arial', 14), bd=5)
        lbl_contact.grid(row=5, sticky=W)

        #===================ENTRY===============================
        firstname = Entry(ContactForm, textvariable=FIRSTNAME, font=('arial', 14))
        firstname.grid(row=0, column=1)
        lastname = Entry(ContactForm, textvariable=LASTNAME, font=('arial', 14))
        lastname.grid(row=1, column=1)
        RadioGroup.grid(row=2, column=1)
        age = Entry(ContactForm, textvariable=AGE,  font=('arial', 14))
        age.grid(row=3, column=1)
        address = Entry(ContactForm, textvariable=ADDRESS,  font=('arial', 14))
        address.grid(row=4, column=1)
        contact = Entry(ContactForm, textvariable=CONTACT,  font=('arial', 14))
        contact.grid(row=5, column=1)
        

        #==================BUTTONS==============================
        btn_addcon = Button(ContactForm, text="Save", width=50, command=SubmitData)
        btn_addcon.grid(row=6, columnspan=2, pady=10)




        
    #============================FRAMES======================================
    Top = Frame(root, width=500, bd=1, relief=SOLID)
    Top.pack(side=TOP)
    Mid = Frame(root, width=500,  bg="#6666ff")
    Mid.pack(side=TOP)
    MidLeft = Frame(Mid, width=100)
    MidLeft.pack(side=LEFT, pady=10)
    MidLeftPadding = Frame(Mid, width=370, bg="#6666ff")
    MidLeftPadding.pack(side=LEFT)
    MidRight = Frame(Mid, width=100)
    MidRight.pack(side=RIGHT, pady=10)
    TableMargin = Frame(root, width=500)
    TableMargin.pack(side=TOP)
    #============================LABELS======================================
    lbl_title = Label(Top, text="جديد کاربر ثبت", font=('arial', 16), width=500)
    lbl_title.pack(fill=X)

    #============================ENTRY=======================================

    #============================BUTTONS=====================================
    btn_add = Button(MidLeft, text="+ ADD NEW", bg="#66ff66", command=AddNewWindow)
    btn_add.pack()
    btn_delete = Button(MidRight, text="DELETE", bg="red", command=DeleteData)
    btn_delete.pack(side=RIGHT)

    #============================TABLES======================================
    scrollbarx = Scrollbar(TableMargin, orient=HORIZONTAL)
    scrollbary = Scrollbar(TableMargin, orient=VERTICAL)
    tree = ttk.Treeview(TableMargin, columns=("MemberID", "Firstname", "Lastname", "Gender", "Age", "Address", "Contact"), height=400, selectmode="extended", yscrollcommand=scrollbary.set, xscrollcommand=scrollbarx.set)
    scrollbary.config(command=tree.yview)
    scrollbary.pack(side=RIGHT, fill=Y)
    scrollbarx.config(command=tree.xview)
    scrollbarx.pack(side=BOTTOM, fill=X)
    tree.heading('MemberID', text="MemberID", anchor=W)
    tree.heading('Firstname', text="Firstname", anchor=W)
    tree.heading('Lastname', text="Lastname", anchor=W)
    tree.heading('Gender', text="Gender", anchor=W)
    tree.heading('Age', text="Age", anchor=W)
    tree.heading('Address', text="Address", anchor=W)
    tree.heading('Contact', text="Contact", anchor=W)
    tree.column('#0', stretch=NO, minwidth=0, width=0)
    tree.column('#1', stretch=NO, minwidth=0, width=0)
    tree.column('#2', stretch=NO, minwidth=0, width=80)
    tree.column('#3', stretch=NO, minwidth=0, width=120)
    tree.column('#4', stretch=NO, minwidth=0, width=90)
    tree.column('#5', stretch=NO, minwidth=0, width=80)
    tree.column('#6', stretch=NO, minwidth=0, width=180)
    tree.column('#7', stretch=NO, minwidth=0, width=120)
    tree.pack()
    tree.bind('<Double-Button-1>', OnSelected)

    #============================INITIALIZATION==============================
    if __name__ == '__main__':
        Database()
        root.mainloop()
        


    new_user.mainloop()


def Help():
    root = tk.Tk()
    root.title("Help")
    root.geometry("650x450")
    root.configure(bg='white')

    lb1 = Label(root , text="""


    """,bg="white")
    lb1.config(font=("B Nazanin",12))
    lb1.place(x=20,y=80)

    text1 = tk.Text(root, height=20, width=30)
    photo = tk.PhotoImage(file="help-icon-26.png")
    text1.insert(tk.END, '\n')
    text1.image_create(tk.END, image=photo)
    text1.pack(side=tk.LEFT)
    text2 = tk.Text(root, height=20, width=50)
    scroll = tk.Scrollbar(root, command=text2.yview)
    text2.configure(yscrollcommand=scroll.set)
    text2.tag_configure('bold_italics', font=('Times New Roman', 9, 'bold', 'italic'))
    text2.tag_configure('big', font=('Times New Roman', 12, 'bold'))
    text2.tag_configure('color',
                        foreground='#476042',
                        font=('Times New Roman', 12, 'bold'))
    text2.tag_bind('follow',
                   '<1>',
                   lambda e, t=text2: t.insert(tk.END, "Not now, maybe later!"))
    text2.insert(tk.END,'\n    Diagnosis of Diabetes using Artificial Intelligence\n', 'big')
    quote = """
    Artificial intelligence/Machine learning (AI/ML):
    is transforming all spheres of our life,
    including the healthcare system. Application of AI/ML ,
    has a potential to vastly enhance the reach of diabetes
    care thereby making it more efficient.
    Thanks: Ardabil University of Medical Sciences
    """
    text2.insert(tk.END, quote, 'color')
    text2.insert(tk.END, 'follow-up\n', 'follow')
    text2.pack(side=tk.LEFT)
    scroll.pack(side=tk.RIGHT, fill=tk.Y)



    

        #logo
    #logo = PhotoImage(file="help-icon-26.png")
    #logolbl = Label(root,image= logo,text="",bg='white')
    #logolbl.place(x=20,y=15)
    
    
        #Home FAQ ContactUs Help
    def ret():
        root.destroy()
        home()
    hom = Button(root,text="Home",bg='white',relief=FLAT,command=ret,cursor='hand2')#,command=Homepage)
    hom.config(font=("Times New Roman",12))
    hom.place(x=350,y=27)
    def ret1():
        root.destroy()
        Faq()
    faq = Button(root,text="FAQ",bg='white',relief=FLAT,cursor='hand2',command=ret1)
    faq.config(font=("Times New Roman",12))
    faq.place(x=525,y=27)
    def ret2():
        root.destroy()
        contactus()
    con = Button(root,text="Contact Us",bg='white',relief=FLAT,command=ret2,cursor='hand2')
    con.config(font=("Times New Roman",12))
    con.place(x=410,y=27)

    hel = Button(root,text="Help",bg='white',relief=FLAT) #command=search
    hel.config(font=("Times New Roman",12))
    hel.place(x=580,y=27)



    
    
   
    #con = PhotoImage(file="icons8-diabetes-16.png")
    #conlbl = Label(root,image= con,text="",bg='white')
    #conlbl.place(x=-2,y=90)
    

    #foot = PhotoImage(file="logo1.png")
    #footlb = Label(root,image= foot,text="",bg='white')
    #footlb.place(x=-2,y=330)


    root.mainloop()

def Faq():
    root = tk.Tk()
    root.title("FAQ's")
    root.geometry("650x450")
    root.configure(bg='white')

        #logo
    #logo = PhotoImage(file="icons8-diabetes-50.png")
    #logolbl = Label(root,image= logo,text="",bg='white')
    #logolbl.place(x=20,y=15)
    
        #Home FAQ ContactUs Help
    def ret():
        root.destroy()
        home()
    hom = Button(root,text="Home",bg='white',relief=FLAT,command=ret,cursor='hand2')#,command=Homepage)
    hom.config(font=("Times New Roman",12))
    hom.place(x=350,y=27)

    faq = Button(root,text="FAQ",bg='white',relief=FLAT)
    faq.config(font=("Times New Roman",12))
    faq.place(x=525,y=27)
    def ret2():
        root.destroy()
        contactus()
    con = Button(root,text="Contact Us",bg='white',relief=FLAT,cursor='hand2',command=ret2)
    con.config(font=("Times New Roman",12))
    con.place(x=410,y=27)
    def ret1():
        root.destroy()
        Help()
    hel = Button(root,text="Help",bg='white',relief=FLAT,command=ret1,cursor='hand2') #command=search
    hel.config(font=("Times New Roman",12))
    hel.place(x=580,y=27)
    
   
    con = PhotoImage(file="faq.png")
    conlbl = Label(root,image= con,text="",bg='white')
    conlbl.place(x=-2,y=90)
    

    #foot = PhotoImage(file="C:/Users/USER/Desktop/ML Project/Photos/footer2.png")
    #footlb = Label(root,image= foot,text="",bg='white')
    #footlb.place(x=-2,y=330)


    root.mainloop()
def contactus():
    root = tk.Tk()
    root.title("Contact Us")
    root.geometry("650x450")
    root.configure(bg='white')

        #logo
    logo = PhotoImage(file="contact-us-png.png")
    logolbl = Label(root,image= logo,text="",bg='white')
    logolbl.place(x=20,y=15)
    
        #Home FAQ ContactUs Help
    def ret():
        root.destroy()
        home()
    hom = Button(root,text="Home",bg='white',relief=FLAT,command=ret,cursor='hand2')#,command=Homepage)
    hom.config(font=("Times New Roman",12))
    hom.place(x=350,y=27)
    def ret1():
        root.destroy()
        Faq()
    faq = Button(root,text="FAQ",bg='white',relief=FLAT,command=ret1,cursor='hand2')
    faq.config(font=("Times New Roman",12))
    faq.place(x=525,y=27)

    con = Button(root,text="Contact Us",bg='white',relief=FLAT)
    con.config(font=("Times New Roman",12))
    con.place(x=410,y=27)
    def ret2():
        root.destroy()
        Help()
    hel = Button(root,text="Help",bg='white',relief=FLAT,cursor='hand2',command=ret2) #command=search
    hel.config(font=("Times New Roman",12))
    hel.place(x=580,y=27)
    
   
    #con = PhotoImage(file="C:/Users/USER/Desktop/ML Project/Photos/contactus.png")
    #conlbl = Label(root,image= con,text="",bg='white')
    #conlbl.place(x=-2,y=90)
    

    #foot = PhotoImage(file="C:/Users/USER/Desktop/ML Project/Photos/footer2.png")
    #footlb = Label(root,image= foot,text="",bg='white')
    #footlb.place(x=-2,y=330)


    root.mainloop()

def home():
    #welcome()
    #Home page
    window = Tk()
    window.title("Diabetes.AI")
    window.geometry("650x450")
    window.configure(bg='white')


    

    #logo
    logo = PhotoImage(file="diabetes.png")
    logolbl = Label(window,image= logo,text="",bg='white')
    logolbl.place(x=20,y=15)

    #Home FAQ ContactUs Help
    hom = Button(window,text="Home",bg='white',relief=FLAT) #command=search
    hom.config(font=("Times New Roman",12))
    hom.place(x=350,y=27)
    
    def fun2():
        window.destroy()
        Faq()
    faq = Button(window,text="FAQ",bg='white',relief=FLAT,cursor='hand2',command=fun2) #command=search
    faq.config(font=("Times New Roman",12))
    faq.place(x=525,y=27)
    
    def fun1():
        window.destroy()
        contactus()
    con = Button(window,text="Contact Us",bg='white',relief=FLAT,command=fun1,cursor='hand2') #command=search
    con.config(font=("Times New Roman",12))
    con.place(x=410,y=27)
    def fun3():
        window.destroy()
        Help()
    hel = Button(window,text="Help",bg='white',relief=FLAT,cursor='hand2',command=fun3) #command=search
    hel.config(font=("Times New Roman",12))
    hel.place(x=580,y=27)

    Filing = Button(window,text="Filing",bg='white',relief=FLAT,command=new_user) #command=new_user
    Filing.config(font=("Times New Roman",12))
    Filing.place(x=290,y=27)


    Vital_Signs = Button(window,text="Vital_Signs",bg='white',relief=FLAT,command=Insert) #command=search
    Vital_Signs.config(font=("Times New Roman",12))
    Vital_Signs.place(x=200,y=27)
    


    
    def fun():
        window.destroy()
        Health()
        
    #Health
    #health = PhotoImage(file="C:/Users/USER/Desktop/ML Project/Photos/health2.png")
    #healthlb = Label(window,image= health,text="",bg='white')
    #healthlb.place(x = 450, y = 115)
    healthbtn = Button(window,text="Health Care",bg='white',relief=FLAT,cursor="hand2",command=fun) 
    healthbtn.config(font=("Edwardian Script ITC",40))
    healthbtn.place(x=222,y=172)

    #healthbtn = Button(window,text="Artificial Intelligence to diagnosis of Diabetes",bg='white',relief=FLAT,cursor="hand2") 
    #healthbtn.config(font=("Edwardian Script ITC",32))
    #healthbtn.place(x=40,y=240)

    #var = StringVar()
    #label = Label( window, textvariable=var)

    #var.set("Artificial Intelligence to diagnosis of Diabetes", font=("Helvetica", 16))
    #label.place(x=190,y=280)
    w = Label(window, text="Artificial Intelligence to diagnosis of Diabetes", font=("Edwardian Script ITC", 25))
    w.place(x=130,y=250)

    #explanation = """At present, only GIF and PPM/PGM"""
    #w2 = tk.Label(window, justify=tk.LEFT,padx = 10,text=explanation).pack(side="left")
    #footer
    #foot = PhotoImage(file="healhcare.png")
    #footlb = Label(window,image= foot,text="",bg='white')
    #footlb.place(x=-2,y=330)

    menu = Menu(window) 
    window.config(menu=menu) 
    filemenu = Menu(menu) 
    menu.add_cascade(label='File', menu=filemenu) 
    filemenu.add_command(label='New') 
    filemenu.add_command(label='Open...') 
    filemenu.add_separator() 
    filemenu.add_command(label='Exit', command=window.quit) 
    helpmenu = Menu(menu) 
    menu.add_cascade(label='Help', menu=helpmenu) 
    helpmenu.add_command(label='About') 

    window.mainloop()
def welcome():
    #loading page
    def task():
        # The window will stay open until this function call ends.
        sleep(2) # Replace this with the code you want to run
        root.destroy()

    root = tk.Tk()
    root.title("Welcome")
    root.geometry("300x250")
    root.configure(bg='white')
    load = PhotoImage(file="images.png")
    loadlbl = tk.Label(root,image= load,text="",bg='white')
    loadlbl.place(x=30,y=40)

    root.after(400, task)
    root.mainloop()
    home()

def Health():
    root = tk.Tk()
    root.title("Diabete.AI ( Health Care ) ")
    root.geometry("650x450")
    root.configure(bg='white')

        #logo
    logo = PhotoImage(file="diabetes2.png")
    logolbl = Label(root,image= logo,text="",bg='white')
    logolbl.place(x=130,y=150)

    
    def ret():
        root.destroy()
        home()
        #Home FAQ ContactUs Help
    hom = Button(root,text="Home",bg='white',relief=FLAT,command=ret,cursor='hand2') #command=Homepage
    hom.config(font=("Times New Roman",12))
    hom.place(x=350,y=27)
    def fun2():
        root.destroy()
        Faq()
    faq = Button(root,text="FAQ",bg='white',relief=FLAT,command=fun2,cursor='hand2')
    faq.config(font=("Times New Roman",12))
    faq.place(x=525,y=27)
    def fun3():
        root.destroy()
        contactus()
    con = Button(root,text="Contact Us",bg='white',relief=FLAT,command=fun3,cursor='hand2')
    con.config(font=("Times New Roman",12))
    con.place(x=410,y=27)
    def fun4():
        root.destroy()
        Help()
    hel = Button(root,text="Help",bg='white',relief=FLAT,command=fun4,cursor='hand2') #command=search
    hel.config(font=("Times New Roman",12))
    hel.place(x=580,y=27)
    
    #filtering
    lb1 = Label(root,text="Select disease",bg="white")
    lb1.config(font=("Times New Roman",14))
    lb1.place(x=20,y=110)

    
    #female diab menu
    def Fdiab():
        #submit button
        def submit(data):
            a,b,c,d,e,f,g,h=eval(data[0]),eval(data[1]),eval(data[2]),eval(data[3]),eval(data[4]),eval(data[5]),eval(data[6]),eval(data[7])
            J='parham_Abdollahi'
            print("This model uses Artificial Inteligence .\nAccuracy of model: 100%\nWe have used PIMA Indians diabetes dataset from UCI archive.")
            print("Patient's Details")
            table = {'Pregancies': a, 'Glocose': b, 'Blood Pressure': c , 'Skine ThickNess': d, 'Insuline': e, 'BMI': f , 'Pedigree Function': g, 'Age': h}
            for name, phone in table.items():
                print(f'{name:10} ==> {phone:10d}')


            # Create DataSet For New patients
            wb = load_workbook('DataSetDiabeticFamle.xlsx') 
            # create the sheet object 
            sheet = wb.active
            sheet.column_dimensions['A'].width = 10
            sheet.column_dimensions['B'].width = 10
            sheet.column_dimensions['C'].width = 10
            sheet.column_dimensions['D'].width = 10
            sheet.column_dimensions['E'].width = 10
            sheet.column_dimensions['F'].width = 10
            sheet.column_dimensions['G'].width = 10
            sheet.column_dimensions['H'].width = 10
            sheet.column_dimensions['I'].width = 10
            #Define New User
            sheet.column_dimensions['J'].width = 50

            # write given data to an excel spreadsheet 
            # at particular location 
            #sheet.cell(row=1, column=1).value = "NameUser"
            sheet.cell(row=1, column=1).value = "Pregnancies"
            sheet.cell(row=1, column=2).value = "Glucose"
            sheet.cell(row=1, column=3).value = "BloodPressure"
            sheet.cell(row=1, column=4).value = "SkinThickness"
            sheet.cell(row=1, column=5).value = "Insulin"
            sheet.cell(row=1, column=6).value = "BMI"
            sheet.cell(row=1, column=7).value = "DiabetesPedigreeFunction"
            sheet.cell(row=1, column=8).value = "Age"
            sheet.cell(row=1, column=9).value = "Outcome"
            #Define New User
            sheet.cell(row=1, column=10).value = "NameUser"
            

            current_row = sheet.max_row 
            current_column = sheet.max_column 
          
            # get method returns current text 
            # as string which we write into 
            # excel spreadsheet at particular location 
            sheet.cell(row=current_row + 1, column=1).value = a 
            sheet.cell(row=current_row + 1, column=2).value = b
            sheet.cell(row=current_row + 1, column=3).value = c
            sheet.cell(row=current_row + 1, column=4).value = d
            sheet.cell(row=current_row + 1, column=5).value = e
            sheet.cell(row=current_row + 1, column=6).value = f
            sheet.cell(row=current_row + 1, column=7).value = g
            sheet.cell(row=current_row + 1, column=8).value = h
            sheet.cell(row=current_row + 1, column=10).value = J
            
            #Define New User
            
            
 
        



            pred = rfc.predict([[a,b,c,d,e,f,g,h]])
            lookup_Diabetic_name_rfc=[pred[0]]
            print('[1]','RFC=',lookup_Diabetic_name_rfc)

            pred1 = MLP.predict([[a,b,c,d,e,f,g,h]])
            lookup_Diabetic_name_MLP=[pred1[0]]
            print('[2]','RFC=',lookup_Diabetic_name_MLP)
            
            pred2 = rfc.predict([[a,b,c,d,e,f,g,h]])
            lookup_Diabetic_name_knn=[pred2[0]]
            print('[3]','knn=',lookup_Diabetic_name_knn)



            pred3 = rfc.predict([[a,b,c,d,e,f,g,h]])
            lookup_Diabetic_name_ABC=[pred3[0]]
            print('[4]','AdaBoost=',lookup_Diabetic_name_ABC)


            pred4 = rfc.predict([[a,b,c,d,e,f,g,h]])
            lookup_Diabetic_name_TREE=[pred4[0]]
            print('[5]','Tree=',lookup_Diabetic_name_TREE)


            pred5 = rfc.predict([[a,b,c,d,e,f,g,h]])
            lookup_Diabetic_name_GNB=[pred5[0]]
            print('[6]','GaussianNB=',lookup_Diabetic_name_GNB)



            pred6 = rfc.predict([[a,b,c,d,e,f,g,h]])
            lookup_Diabetic_name_GBC=[pred6[0]]
            print('[7]','GradientBoosting=',lookup_Diabetic_name_GBC)



            pred7 = rfc.predict([[a,b,c,d,e,f,g,h]])
            lookup_Diabetic_name_SVM=[pred7[0]]
            print('[8]','SVM=',lookup_Diabetic_name_SVM)

            pred8 = rfc.predict([[a,b,c,d,e,f,g,h]])
            lookup_Diabetic_name_ETC=[pred8[0]]
            print('[9]','ETC=',lookup_Diabetic_name_ETC)

            predstack=sl.predict([[a,b,c,d,e,f,g,h]])
            lookup_Diabetic_name_Stack=[predstack[0]]
            print('[10]','Stack=',lookup_Diabetic_name_Stack)
            r = int("".join(map(str, lookup_Diabetic_name_Stack)))

            
            sheet.cell(row=current_row + 1, column=9).value = r
            # save the file 
            wb.save('DataSetDiabeticFamle.xlsx') 



    
            

            
            #Accuracy for Predict Algorithms
            predicted=rfc.predict(X_test)
            accurace_rfc=np.mean(predicted==y_test)
            print('[1]','accuracy for RandomForestClassifier =%.2f'%accurace_rfc)


            predicted=knn.predict(X_test)
            accurace_knn=np.mean(predicted==y_test)
            print('[2]','accuracy for KNeighborsClassifier =%.2f'%accurace_knn)

            predicted=MLP.predict(X_test)
            accurace_MLP=np.mean(predicted==y_test)
            print('[3]','accuracy for KNeighborsClassifier =%.2f'%accurace_MLP)

    
        
            predicted=ABC.predict(X_test)
            accurace_ABC=np.mean(predicted==y_test)
            print('[4]','accuracy for AdaBoostClassifier =%.2f'%accurace_ABC)
        
        
            predicted=DTree.predict(X_test)
            accurace_Dtree=np.mean(predicted==y_test)
            print('[5]','accuracy for tree =%.2f'%accurace_Dtree)
        
        
            predicted=GNB.predict(X_test)
            accurace_GNB=np.mean(predicted==y_test)
            print('[6]','accuracy for GaussianNB =%.2f'%accurace_GNB)
        
        
            predicted=gbc.predict(X_test)
            accurace_gbc=np.mean(predicted==y_test)
            print('[7]','accuracy for GradientBoostingClassifier =%.2f'%accurace_gbc)
        
            predicted=SVM.predict(X_test)
            accurace_SVM=np.mean(predicted==y_test)
            print('[8]','accuracy for Support Vector Machine =%.2f'%accurace_SVM)
        
            predicted=ETC.predict(X_test)
            accurace_ETC=np.mean(predicted==y_test)
            print('[9]','accuracy for ExtraTreesClassifier =%.2f'%accurace_ETC)
            #score = metrics.accuracy_score(y1, ETC.predict(x1))
            #print("Accuracy: %f" % score)

            predicted=sl.predict(X_test)
            accurace_Stack=np.mean(predicted==y_test)
            print('[10]','accuracy for StackedClassifier =%.2f'%accurace_Stack)
            score = metrics.accuracy_score(y1, sl.predict(x1))
            print("Accuracy: %f" % score)

            print('_______________________________')

            a=lookup_Diabetic_name_rfc
            b=lookup_Diabetic_name_MLP
            c=lookup_Diabetic_name_knn
            d=lookup_Diabetic_name_ABC
            e=lookup_Diabetic_name_TREE
            f=lookup_Diabetic_name_GNB
            g=lookup_Diabetic_name_GBC
            h=lookup_Diabetic_name_SVM
            i=lookup_Diabetic_name_ETC
            j=lookup_Diabetic_name_Stack
            resultD=(a+b+c+d+e+f+g+h+i+j)
            print('lookup_Diabetic=',resultD)

            import collections
            prediict=collections.Counter(resultD)
            Nagative=collections.Counter(prediict)[0]
            Positive=collections.Counter(prediict)[1]
            print('Nagative',collections.Counter(prediict)[0])
            print('Positive',collections.Counter(prediict)[1])
            print('Accuracy : ',accurace_Stack)
            print(prediict)
            
            W=(collections.Counter(prediict)[0] * 100)
            E=(collections.Counter(prediict)[1] * 100)
            X=(W / E)
            print("Diabetes diagnosis percentage : " , X * 10)

            

            print('_________________THE END______________________')
                    #Staking_Classifier
            
            from tkinter import messagebox

            messagebox.showinfo("score", score)
            messagebox.showinfo("prediict", prediict)
            
            pred=lookup_Diabetic_name_rfc
            RFC=accurace_rfc

            pred1=lookup_Diabetic_name_MLP
            MLP1=accurace_MLP

            pred2=lookup_Diabetic_name_knn
            KNN1=accurace_knn

              
            pred3=lookup_Diabetic_name_ABC
            ABC1=accurace_ABC
              
            pred4=lookup_Diabetic_name_TREE
            TREE1=accurace_Dtree
              
            pred5=lookup_Diabetic_name_GNB
            GNB1=accurace_GNB

            pred6=lookup_Diabetic_name_GBC
            GBC1=accurace_gbc
              
            pred7=lookup_Diabetic_name_SVM
            SVM1=accurace_SVM
              
            pred8=lookup_Diabetic_name_ETC
            ETC1=accurace_ETC

            predstack=lookup_Diabetic_name_Stack
            Stack1=accurace_Stack

            if  r == 0 :
                print("Diagnosis Accurately detect suggests  {} that patient does not suffers from diabetes. {}.".format(Stack1,predstack))
                messagebox.showwarning("Result", "Diagnosis suggests that patient does not suffers from diabetes.")
                
            else:
                print("Our diagnosis  Accurately detect {} suggests patient does suffer from diabetes {}. \nPlease get checked soon. ".format(Stack1,predstack))
                messagebox.showwarning("Result", "Our diagnosis suggests patient does suffer from diabetes.\nPlease get checked soon.")
            

            # Print Result In New Form #
            root2 = Toplevel()
            root2.title("Result ( Diabetes type2 ) ")
            root2.geometry("350x450")
            root2.configure(bg='white')
            Pregancies = {'Pregancies': data[0]}
            Glocose = { 'Glocose': data[1]}
            BloodPressure = {'Blood Pressure': data[2]}
            Skine = {'Skine ThickNess': data[3]}
            Insuline = {'Insuline': data[4]}
            BMI = {'BMI': data[5]}
            Pedigree = {'Pedigree Function': data[6]}
            Age = {'Age': data[7]}

            #print("First Name: %s\nLast Name: %s" % (a.get(), b.get()))
            #a.delete(0, tk.END)
            #b.delete(0, tk.END)
            
            


            w = Label(root2, text="Patient's Details")
            w.pack()
            w = Label(root2, text=Pregancies)
            w.pack()
            w = Label(root2, text=Glocose)
            w.pack()
            w = Label(root2, text=BloodPressure)
            w.pack()
            w = Label(root2, text=Insuline)
            w.pack()
            w = Label(root2, text=BMI)
            w.pack()
            w = Label(root2, text=Pedigree)
            w.pack()
            w = Label(root2, text=Age)
            w.pack()
            
            

            w = Label(root2, text="Negative[0]")
            w.pack()
            w = Label(root2, text=collections.Counter(prediict)[0])
            w.pack()
            w = Label(root2, text="Positive[1]")
            w.pack()
            w = Label(root2, text=collections.Counter(prediict)[1])
            w.pack()
            w = Label(root2, text="Accuracy")
            w.pack()
            w = Label(root2, text=accurace_Stack)
            w.pack()
            w = Label(root2, text="Predict")
            w.pack()
            w = Label(root2, text=prediict)
            w.pack()

            if r == 0 :
                w = Label(root2, text="Result : Diagnosis suggests that patient does not suffers from diabetes.")
                w.pack()
            else:
                w = Label(root2, text="Result :" "Our diagnosis suggests patient does suffer from diabetes.\nPlease get checked soon.")
                w.pack()

            w = Label(root2, text="This model uses Artificial Inteligence .\nAccuracy of model: 100%\nWe have used PIMA Indians diabetes dataset from UCI archive.")
            w.pack()

            # Print Result In New Form #


            
            
            
            with open('female.txt', 'w') as a:
                #a.writelines(user)
                a.write("\n")
                a.write('RFC Predict =')
                a.write('%s:' % pred)
                a.write("\n")
                a.write(' Accuracy RFC =')
                a.write('%s:' % RFC)
                a.write("\n")
                a.write('_________________')

                a.write("\n")
                a.write('MLP Predict =')
                a.write('%s:' % pred1)
                a.write("\n")
                a.write(' Accuracy MLP =')
                a.write('%s:' % MLP1)
                a.write("\n")
                a.write('_________________')

                a.write("\n")
                a.write('KNN Predict =')
                a.write('%s:' % pred2)
                a.write("\n")
                a.write(' Accuracy KNN =')
                a.write('%s:' % KNN1)
                a.write("\n")
                a.write('_________________')




                a.write("\n")
                a.write('ABC Predict =')
                a.write('%s:' % pred3)
                a.write("\n")
                a.write(' Accuracy ABC =')
                a.write('%s:' % ABC1)
                a.write("\n")
                a.write('_________________')




                a.write("\n")
                a.write('TREE Predict =')
                a.write('%s:' % pred4)
                a.write("\n")
                a.write(' Accuracy TREE =')
                a.write('%s:' % TREE1)
                a.write("\n")
                a.write('_________________')

                a.write("\n")
                a.write('GNB Predict =')
                a.write('%s:' % pred5)
                a.write("\n")
                a.write(' Accuracy GNB =')
                a.write('%s:' % GNB1)
                a.write("\n")
                a.write('_________________')


                a.write("\n")
                a.write('GBC Predict =')
                a.write('%s:' % pred6)
                a.write("\n")
                a.write(' Accuracy GBC =')
                a.write('%s:' % GBC1)
                a.write("\n")
                a.write('_________________')



                a.write("\n")
                a.write('SVM Predict =')
                a.write('%s:' % pred7)
                a.write("\n")
                a.write(' Accuracy SVM =')
                a.write('%s:' % SVM1)
                a.write("\n")
                a.write('_________________')



                a.write("\n")
                a.write('ETC Predict =')
                a.write('%s:' % pred8)
                a.write("\n")
                a.write(' Accuracy ETC =')
                a.write('%s:' % ETC1)
                a.write("\n")
                a.write('_________________')

                a.write("\n")
                a.write('Stack Predict =')
                a.write('%s:' % predstack)
                a.write("\n")
                a.write(' Accuracy Stack =')
                a.write('%s:' % Stack1)
                a.write("\n")
                a.write('_________________')

                

                #Result
                a.write("\n")
                a.write('Nagative =')
                a.write('%s:' % collections.Counter(prediict)[0])
                a.write("\n")
                a.write(' Positive =')
                a.write('%s:' % collections.Counter(prediict)[1])
                a.write("\n")
                a.write(' Diabetes diagnosis percentage =')
                a.write('%s:' % accurace_Stack)
                a.write("\n")
                a.write('________Result_________')
                a.write("\n")


                if  r == 0 :                   
                    print("Diagnosis Accurately detect suggests  {} that patient does not suffers from diabetes. {}.".format(Stack1,predstack))
                    a.write("Diagnosis suggests that patient does not suffers from diabetes.")
                else:
                    print("Our diagnosis  Accurately detect {} suggests patient does suffer from diabetes {}. \nPlease get checked soon. ".format(Stack1,predstack))
                    a.write("Our diagnosis suggests patient does suffer from diabetes.\nPlease get checked soon.")
                    

                    


                #Result


                




            
            return pred
            return pred1
            return pred2
            return pred3
            return pred4
            return pred5
            return pred6
            return pred7
            return pred8
            return predstack
        
        def check():
            if submit([text1.get(),text2.get(),text3.get(),text4.get(),text5.get(),text6.get(),text7.get(),text8.get()]):
                root = Toplevel()
                root.title("Result ( Diabetes type2 ) ")
                root.geometry("350x250")
                root.configure(bg='white')

                
                
                #logo
                #logo = PhotoImage(file="C:/Users/USER/Desktop/ML Project/Photos/logo1.png")
                #logolbl = Label(root,image= logo,text="",bg='white')
                #logolbl.place(x=20,y=15)
                def fun3():
                    root.destroy()
                back = Button(root,text="<<<",bg='White',relief=FLAT,command=fun3,cursor='hand2')
                back.config(font=("Times New Roman",18))
                back.place(x=410,y=24)

                

                
                
                
                #lb1 = PhotoImage(file="C:/Users/USER/Desktop/ML Project/Photos/havediabetes.png")
                #lb2 = Label(root,image= lb1,text="",bg='white')
                #lb2.place(x=-2,y=90)
                #footer
                #foot = PhotoImage(file="C:/Users/USER/Desktop/ML Project/Photos/footer2.png")
                #footlb = Label(root,image= foot,text="",bg='white')
                #footlb.place(x=-2,y=330)
                root.mainloop()
            else:
                root = Toplevel()
                root.title("Result ( Diabetes type2 ) ")
                root.geometry("650x450")
                root.configure(bg='white')
                #logo
                #logo = PhotoImage(file="C:/Users/USER/Desktop/ML Project/Photos/logo1.png")
                #logolbl = Label(root,image= logo,text="",bg='white')
                #logolbl.place(x=20,y=15)
                def fun3():
                    root.destroy()
                back = Button(root,text="<<<",bg='white',relief=FLAT,command=fun3,cursor='hand2')
                back.config(font=("Times New Roman",18))
                back.place(x=410,y=24)

                

                

                
                
                
                #lb1 = PhotoImage(file="C:/Users/USER/Desktop/ML Project/Photos/nodiabetes.png")
                #lb2 = Label(root,image= lb1,text="",bg='white')
                #lb2.place(x=-2,y=90)
                #footer
                #foot = PhotoImage(file="C:/Users/USER/Desktop/ML Project/Photos/footer2.png")
                #footlb = Label(root,image= foot,text="",bg='white')
                #footlb.place(x=-2,y=330)
                root.mainloop()
        #1
        lb1 = Label(root,text="Pregnancies    :",bg="white")
        lb1.config(font=("Times New Roman",12))
        lb1.place(x=20,y=190)
        text1 = tk.Entry(root,width=10,relief='solid',borderwidth=1,bg='snow')
        text1.place(x=185,y=193)
        #2
        lb1 = Label(root,text="Glucose        :",bg="white")
        lb1.config(font=("Times New Roman",12))
        lb1.place(x=20,y=220)
        lb1 = Label(root,text="70-180 mg/dl",bg="white")
        lb1.config(font=("Times New Roman",9))
        lb1.place(x=250,y=220)
        text2 = tk.Entry(root,width=10,relief='solid',borderwidth=1,bg='snow')
        text2.place(x=185,y=223)
        #3
        lb1 = Label(root,text="Blood Pressure :",bg="white")
        lb1.config(font=("Times New Roman",12))
        lb1.place(x=20,y=250)
        lb1 = Label(root,text="89-140mm HG",bg="white")
        lb1.config(font=("Times New Roman",9))
        lb1.place(x=250,y=250)
        text3 = tk.Entry(root,width=10,relief='solid',borderwidth=1,bg='snow')
        text3.place(x=185,y=253)
        #4
        lb1 = Label(root,text="Skin Thickness :",bg="white")
        lb1.config(font=("Times New Roman",12))
        lb1.place(x=20,y=280)
        lb1 = Label(root,text="10-50mm",bg="white")
        lb1.config(font=("Times New Roman",9))
        lb1.place(x=250,y=280)
        text4 = tk.Entry(root,width=10,relief='solid',borderwidth=1,bg='snow')
        text4.place(x=185,y=283)
        #5
        lb1 = Label(root,text="Insulin           :",bg="white")
        lb1.config(font=("Times New Roman",12))
        lb1.place(x=340,y=190)
        lb1 = Label(root,text="15-276mu U/mL",bg="white")
        lb1.config(font=("Times New Roman",9))
        lb1.place(x=570,y=190)
        text5 = tk.Entry(root,width=10,relief='solid',borderwidth=1,bg='snow')
        text5.place(x=505,y=193)
        #6
        lb1 = Label(root,text="BMI               :",bg="white")
        lb1.config(font=("Times New Roman",12))
        lb1.place(x=340,y=220)
        lb1 = Label(root,text="10-50",bg="white")
        lb1.config(font=("Times New Roman",9))
        lb1.place(x=570,y=220)
        text6 = tk.Entry(root,width=10,relief='solid',borderwidth=1,bg='snow')
        text6.place(x=505,y=223)
        #7
        lb1 = Label(root,text="Pedigree Function :",bg="white")
        lb1.config(font=("Times New Roman",12))
        lb1.place(x=340,y=250)
        text7 = tk.Entry(root,width=10,relief='solid',borderwidth=1,bg='snow')
        text7.place(x=505,y=253)
        #8
        lb1 = Label(root,text="Age               :",bg="white")
        lb1.config(font=("Times New Roman",12))
        lb1.place(x=340,y=280)
        lb1 = Label(root,text="30-80",bg="white")
        lb1.config(font=("Times New Roman",9))
        lb1.place(x=570,y=280)
        text8 = tk.Entry(root,width=10,relief='solid',borderwidth=1,bg='snow')
        text8.place(x=505,y=283)

        lb1 = Label(root,text="Name User:",bg="white")
        lb1.config(font=("Times New Roman",12))
        lb1.place(x=100,y=340)
        text9 = tk.Entry(root,width=30,relief='solid',borderwidth=1,bg='snow')
        text9.place(x=200,y=340)
        #check button
        chkbtn = Button(root,text="Check Result",relief=FLAT,bg='orange',width=14,command=check,cursor='hand2')
        chkbtn.config(font=("Times New Roman",10))
        chkbtn.place(x=485,y=120)
        
    #male diab menu
    def Mdiab():
        #submit button
        def submit(data):
            a = 0
            b,c,d,e,f,g,h=eval(data[1]),eval(data[2]),eval(data[3]),eval(data[4]),eval(data[5]),eval(data[6]),eval(data[7])
            J='parham_Abdollahi'
            print("This model uses Artificial Inteligence .\nAccuracy of model: 100%\nWe have used PIMA Indians diabetes dataset from UCI archive.")
            print("Patient's Details")
            table = {'Pregancies': a, 'Glocose': b, 'Blood Pressure': c , 'Skine ThickNess': d, 'Insuline': e, 'BMI': f , 'Pedigree Function': g, 'Age': h}
            for name, phone in table.items():
                print(f'{name:10} ==> {phone:10d}')
            
            
            # Create DataSet For New patients
            wb = load_workbook('DataSetDiabeticMale.xlsx') 
            # create the sheet object 
            sheet = wb.active
            sheet.column_dimensions['A'].width = 30
            sheet.column_dimensions['B'].width = 10
            sheet.column_dimensions['C'].width = 10
            sheet.column_dimensions['D'].width = 20
            sheet.column_dimensions['E'].width = 20
            sheet.column_dimensions['F'].width = 40
            sheet.column_dimensions['G'].width = 50
            sheet.column_dimensions['H'].width = 50
            sheet.column_dimensions['I'].width = 50
            sheet.column_dimensions['J'].width = 50

            # write given data to an excel spreadsheet 
            # at particular location 
            sheet.cell(row=1, column=1).value = "Pregnancies"
            sheet.cell(row=1, column=2).value = "Glucose"
            sheet.cell(row=1, column=3).value = "BloodPressure"
            sheet.cell(row=1, column=4).value = "SkinThickness"
            sheet.cell(row=1, column=5).value = "Insulin"
            sheet.cell(row=1, column=6).value = "BMI"
            sheet.cell(row=1, column=7).value = "DiabetesPedigreeFunction"
            sheet.cell(row=1, column=8).value = "Age"
            sheet.cell(row=1, column=9).value = "Outcome"
            sheet.cell(row=1, column=10).value = "NameUser"

            current_row = sheet.max_row 
            current_column = sheet.max_column 
          
            # get method returns current text 
            # as string which we write into 
            # excel spreadsheet at particular location 
            sheet.cell(row=current_row + 1, column=1).value = a 
            sheet.cell(row=current_row + 1, column=2).value = b
            sheet.cell(row=current_row + 1, column=3).value = c
            sheet.cell(row=current_row + 1, column=4).value = d
            sheet.cell(row=current_row + 1, column=5).value = e
            sheet.cell(row=current_row + 1, column=6).value = f
            sheet.cell(row=current_row + 1, column=7).value = g
            sheet.cell(row=current_row + 1, column=8).value = h
            sheet.cell(row=current_row + 1, column=10).value = J





            pred9 = rfc.predict([[a,b,c,d,e,f,g,h]])
            lookup_Diabetic_name_rfc=[pred9[0]]
            print('[1]','RFC=',lookup_Diabetic_name_rfc)

            pred10 = MLP.predict([[a,b,c,d,e,f,g,h]])
            lookup_Diabetic_name_MLP=[pred10[0]]
            print('[2]','RFC=',lookup_Diabetic_name_MLP)
            
            pred11 = rfc.predict([[a,b,c,d,e,f,g,h]])
            lookup_Diabetic_name_knn=[pred11[0]]
            print('[3]','knn=',lookup_Diabetic_name_knn)



            pred12 = rfc.predict([[a,b,c,d,e,f,g,h]])
            lookup_Diabetic_name_ABC=[pred12[0]]
            print('[4]','AdaBoost=',lookup_Diabetic_name_ABC)


            pred13 = rfc.predict([[a,b,c,d,e,f,g,h]])
            lookup_Diabetic_name_TREE=[pred13[0]]
            print('[5]','Tree=',lookup_Diabetic_name_TREE)


            pred14 = rfc.predict([[a,b,c,d,e,f,g,h]])
            lookup_Diabetic_name_GNB=[pred14[0]]
            print('[6]','GaussianNB=',lookup_Diabetic_name_GNB)



            pred15 = rfc.predict([[a,b,c,d,e,f,g,h]])
            lookup_Diabetic_name_GBC=[pred15[0]]
            print('[7]','GradientBoosting=',lookup_Diabetic_name_GBC)



            pred16 = rfc.predict([[a,b,c,d,e,f,g,h]])
            lookup_Diabetic_name_SVM=[pred16[0]]
            print('[8]','SVM=',lookup_Diabetic_name_SVM)

            pred17 = rfc.predict([[a,b,c,d,e,f,g,h]])
            lookup_Diabetic_name_ETC=[pred17[0]]
            print('[9]','ETC=',lookup_Diabetic_name_ETC)

            predstack1=sl.predict([[a,b,c,d,e,f,g,h]])
            lookup_Diabetic_name_Stack=[predstack1[0]]
            print('[10]','Stack=',lookup_Diabetic_name_Stack)

            r = int("".join(map(str, lookup_Diabetic_name_Stack)))

            
            sheet.cell(row=current_row + 1, column=9).value = r
            # save the file 
            wb.save('DataSetDiabeticMale.xlsx')

            
            


            

            #Accuracy for Predicts Algorithms
            predicted=rfc.predict(X_test)
            accurace_rfc=np.mean(predicted==y_test)
            print('[1]','accuracy for RandomForestClassifier =%.2f'%accurace_rfc)


            predicted=knn.predict(X_test)
            accurace_knn=np.mean(predicted==y_test)
            print('[2]','accuracy for KNeighborsClassifier =%.2f'%accurace_knn)

            predicted=MLP.predict(X_test)
            accurace_MLP=np.mean(predicted==y_test)
            print('[3]','accuracy for KNeighborsClassifier =%.2f'%accurace_MLP)

    
        
            predicted=ABC.predict(X_test)
            accurace_ABC=np.mean(predicted==y_test)
            print('[4]','accuracy for AdaBoostClassifier =%.2f'%accurace_ABC)
        
        
            predicted=DTree.predict(X_test)
            accurace_Dtree=np.mean(predicted==y_test)
            print('[5]','accuracy for tree =%.2f'%accurace_Dtree)
        
        
            predicted=GNB.predict(X_test)
            accurace_GNB=np.mean(predicted==y_test)
            print('[6]','accuracy for GaussianNB =%.2f'%accurace_GNB)
        
        
            predicted=gbc.predict(X_test)
            accurace_gbc=np.mean(predicted==y_test)
            print('[7]','accuracy for GradientBoostingClassifier =%.2f'%accurace_gbc)
        
            predicted=SVM.predict(X_test)
            accurace_SVM=np.mean(predicted==y_test)
            print('[8]','accuracy for Support Vector Machine =%.2f'%accurace_SVM)
        
            predicted=ETC.predict(X_test)
            accurace_ETC=np.mean(predicted==y_test)
            print('[9]','accuracy for ExtraTreesClassifier =%.2f'%accurace_ETC)

            predicted=sl.predict(X_test)
            accurace_Stack=np.mean(predicted==y_test)
            print('[10]','accuracy for StackedClassifier =%.2f'%accurace_Stack)
            score = metrics.accuracy_score(y1, sl.predict(x1))

            
            print('_______________________________')

            a=lookup_Diabetic_name_rfc
            b=lookup_Diabetic_name_MLP
            c=lookup_Diabetic_name_knn
            d=lookup_Diabetic_name_ABC
            e=lookup_Diabetic_name_TREE
            f=lookup_Diabetic_name_GNB
            g=lookup_Diabetic_name_GBC
            h=lookup_Diabetic_name_SVM
            i=lookup_Diabetic_name_ETC
            j=lookup_Diabetic_name_Stack
            resultD=(a+b+c+d+e+f+g+h+i+j)
            print('lookup_Diabetic=',resultD)

            import collections
            prediict=collections.Counter(resultD)
            Nagative=collections.Counter(prediict)[0]
            Positive=collections.Counter(prediict)[1]
            print('Nagative',collections.Counter(prediict)[0])
            print('Positive',collections.Counter(prediict)[1])
            print('Accuracy : ',accurace_rfc)
            print(prediict)


            W=(collections.Counter(prediict)[0] * 100)
            E=(collections.Counter(prediict)[1] * 100)
            X=(W / E)
            print("Diabetes diagnosis percentage : " , X * 10)

            print('_________________THE END______________________')


            
             
            


            




            
                
            #print('[Result]','Stack=',lookup_Diabetic_name_Stack)
            
            

            #################################### Print IN Form
            
        

            #lab=Label(window,text=lookup_Diabetic_name_Stack).place(x=210,y=17)
            

            ##############

            from tkinter import messagebox

            messagebox.showinfo("score", score)
            messagebox.showinfo("prediict", prediict)
            
            pred10=lookup_Diabetic_name_rfc
            RFC=accurace_rfc

            pred11=lookup_Diabetic_name_MLP
            MLP1=accurace_MLP

            pred12=lookup_Diabetic_name_knn
            KNN1=accurace_knn

              
            pred13=lookup_Diabetic_name_ABC
            ABC1=accurace_ABC
              
            pred14=lookup_Diabetic_name_TREE
            TREE1=accurace_Dtree
              
            pred15=lookup_Diabetic_name_GNB
            GNB1=accurace_GNB

            pred16=lookup_Diabetic_name_GBC
            GBC1=accurace_gbc
              
            pred17=lookup_Diabetic_name_SVM
            SVM1=accurace_SVM
              
            pred18=lookup_Diabetic_name_ETC
            ETC1=accurace_ETC

            predstack1=lookup_Diabetic_name_Stack
            Stack1=accurace_Stack



            if  r == 0 :
                print("Diagnosis Accurately detect suggests  {} that patient does not suffers from diabetes. {}.".format(Stack1,predstack1))
                messagebox.showwarning("Result", "Diagnosis suggests that patient does not suffers from diabetes.")
                
            else:
                print("Our diagnosis  Accurately detect {} suggests patient does suffer from diabetes {}. \nPlease get checked soon. ".format(Stack1,predstack1))
                messagebox.showwarning("Result", "Our diagnosis suggests patient does suffer from diabetes.\nPlease get checked soon.")
            

            # Print Result In New Form #
            root2 = Toplevel()
            root2.title("Result ( Diabetes type2 ) ")
            root2.geometry("350x450")
            root2.configure(bg='white')
            Pregancies = {'Pregancies': data[0]}
            Glocose = { 'Glocose': data[1]}
            BloodPressure = {'Blood Pressure': data[2]}
            Skine = {'Skine ThickNess': data[3]}
            Insuline = {'Insuline': data[4]}
            BMI = {'BMI': data[5]}
            Pedigree = {'Pedigree Function': data[6]}
            Age = {'Age': data[7]}
            


            w = Label(root2, text="Patient's Details")
            w.pack()
            w = Label(root2, text=Pregancies)
            w.pack()
            w = Label(root2, text=Glocose)
            w.pack()
            w = Label(root2, text=BloodPressure)
            w.pack()
            w = Label(root2, text=Insuline)
            w.pack()
            w = Label(root2, text=BMI)
            w.pack()
            w = Label(root2, text=Pedigree)
            w.pack()
            w = Label(root2, text=Age)
            w.pack()
            
            

            w = Label(root2, text="Negative[0]")
            w.pack()
            w = Label(root2, text=collections.Counter(prediict)[0])
            w.pack()
            w = Label(root2, text="Positive[1]")
            w.pack()
            w = Label(root2, text=collections.Counter(prediict)[1])
            w.pack()
            w = Label(root2, text="Accuracy")
            w.pack()
            w = Label(root2, text=accurace_Stack)
            w.pack()
            w = Label(root2, text="Predict")
            w.pack()
            w = Label(root2, text=prediict)
            w.pack()
            
            if  r == 0 :
                w = Label(root2, text="Result : Diagnosis suggests that patient does not suffers from diabetes.")
                w.pack()
            else:
                w = Label(root2, text="Result :" "Our diagnosis suggests patient does suffer from diabetes.\nPlease get checked soon.")
                w.pack()

            w = Label(root2, text="This model uses Artificial Inteligence .\nAccuracy of model: 100%\nWe have used PIMA Indians diabetes dataset from UCI archive.")
            w.pack()

            
            

    

            
            
            
            with open('male.txt', 'w') as a:
                
                #a.writelines(user)
                a.write("\n")
                a.write('RFC Predict =')
                a.write('%s:' % pred10)
                a.write("\n")
                a.write(' Accuracy RFC =')
                a.write('%s:' % RFC)
                a.write("\n")
                a.write('_________________')

                a.write("\n")
                a.write('MLP Predict =')
                a.write('%s:' % pred11)
                a.write("\n")
                a.write(' Accuracy MLP =')
                a.write('%s:' % MLP1)
                a.write("\n")
                a.write('_________________')

                a.write("\n")
                a.write('KNN Predict =')
                a.write('%s:' % pred12)
                a.write("\n")
                a.write(' Accuracy KNN =')
                a.write('%s:' % KNN1)
                a.write("\n")
                a.write('_________________')




                a.write("\n")
                a.write('ABC Predict =')
                a.write('%s:' % pred13)
                a.write("\n")
                a.write(' Accuracy ABC =')
                a.write('%s:' % ABC1)
                a.write("\n")
                a.write('_________________')




                a.write("\n")
                a.write('TREE Predict =')
                a.write('%s:' % pred14)
                a.write("\n")
                a.write(' Accuracy TREE =')
                a.write('%s:' % TREE1)
                a.write("\n")
                a.write('_________________')

                a.write("\n")
                a.write('GNB Predict =')
                a.write('%s:' % pred15)
                a.write("\n")
                a.write(' Accuracy GNB =')
                a.write('%s:' % GNB1)
                a.write("\n")
                a.write('_________________')


                a.write("\n")
                a.write('GBC Predict =')
                a.write('%s:' % pred16)
                a.write("\n")
                a.write(' Accuracy GBC =')
                a.write('%s:' % GBC1)
                a.write("\n")
                a.write('_________________')



                a.write("\n")
                a.write('SVM Predict =')
                a.write('%s:' % pred17)
                a.write("\n")
                a.write(' Accuracy SVM =')
                a.write('%s:' % SVM1)
                a.write("\n")
                a.write('_________________')



                a.write("\n")
                a.write('ETC Predict =')
                a.write('%s:' % pred18)
                a.write("\n")
                a.write(' Accuracy ETC =')
                a.write('%s:' % ETC1)
                a.write("\n")
                a.write('_________________')



                a.write("\n")
                a.write('Stack Predict =')
                a.write('%s:' % predstack1)
                a.write("\n")
                a.write(' Accuracy Stack =')
                a.write('%s:' % Stack1)
                a.write("\n")
                a.write('_________________')



                #Result
                a.write("\n")
                a.write('Nagative =')
                a.write('%s:' % collections.Counter(prediict)[0])
                a.write("\n")
                a.write(' Positive =')
                a.write('%s:' % collections.Counter(prediict)[1])
                a.write("\n")
                a.write(' Diabetes diagnosis percentage =')
                a.write('%s:' % Stack1)
                a.write("\n")
                a.write('________Result_________')
                a.write("\n")

                if  r == 0 :
                    
                    print("Diagnosis Accurately detect suggests  {} that patient does not suffers from diabetes. {}.".format(Stack1,predstack1))
                    a.write("Diagnosis suggests that patient does not suffers from diabetes.")
                else:
                    print("Our diagnosis  Accurately detect {} suggests patient does suffer from diabetes {}. \nPlease get checked soon. ".format(Stack1,predstack1))
                    a.write("Our diagnosis suggests patient does suffer from diabetes.\nPlease get checked soon.")
                    


                #Result

            
     
            
            
            return pred9
            return pred10
            return pred11
            return pred12
            return pred13
            return pred14
            return pred15
            return pred16
            return predstack1

        def check():
            if submit([text1.get(),text2.get(),text3.get(),text4.get(),text5.get(),text6.get(),text7.get(),text8.get()]):
                root = Toplevel()
                root.title("Predizione ( Health Care ) ")
                root.geometry("650x450")
                root.configure(bg='white')


                #w = Label(root, text="Artificial Intelligence 1", font=("Edwardian Script ITC", 25))
                #w.place(x=130,y=250)
                #logo
                #logo = PhotoImage(file="C:/Users/USER/Desktop/ML Project/Photos/logo1.png")
                #logolbl = Label(root,image= logo,text="",bg='white')
                #logolbl.place(x=20,y=15)
                def fun3():
                    root.destroy()
                back = Button(root,text="<<<",bg='white',relief=FLAT,command=fun3,cursor='hand2')
                back.config(font=("Times New Roman",18))
                back.place(x=410,y=24)
                
                w = Label(root, text="Artificial Intelligence 2", font=("Edwardian Script ITC", 25))
                w.place(x=130,y=250)
                

                
                
                #lb1 = PhotoImage(file="C:/Users/USER/Desktop/ML Project/Photos/havediabetes.png")
                #lb2 = Label(root,image= lb1,text="",bg='white')
                #lb2.place(x=-2,y=90)
                #footer
                #foot = PhotoImage(file="C:/Users/USER/Desktop/ML Project/Photos/footer2.png")
                #footlb = Label(root,image= foot,text="",bg='white')
                #footlb.place(x=-2,y=330)
                root.mainloop()
            else:
                root = Toplevel()
                root.title("Predizione ( Health Care ) ")
                root.geometry("650x450")
                root.configure(bg='white')


                #w = Label(root, text="Artificial Intelligence 3", font=("Edwardian Script ITC", 25))
                #w.place(x=130,y=250) 
                #logo
                #logo = PhotoImage(file="C:/Users/USER/Desktop/ML Project/Photos/logo1.png")
                #logolbl = Label(root,image= logo,text="",bg='white')
                #logolbl.place(x=20,y=15)
                
                def fun3():
                    root.destroy()
                back = Button(root,text="<<<",bg='white',relief=FLAT,command=fun3,cursor='hand2')
                back.config(font=("Times New Roman",18))
                back.place(x=410,y=24)

                #w = Label(root, text="Artificial Intelligence 4", font=("Edwardian Script ITC", 25))
                #w.place(x=130,y=250)
                
                
                #lb1 = PhotoImage(file="C:/Users/USER/Desktop/ML Project/Photos/nodiabetes.png")
                #lb2 = Label(root,image= lb1,text="",bg='white')
                #lb2.place(x=-2,y=90)
                #footer
                #foot = PhotoImage(file="C:/Users/USER/Desktop/ML Project/Photos/footer2.png")
                #footlb = Label(root,image= foot,text="",bg='white')
                #footlb.place(x=-2,y=330)
                root.mainloop()
        #1
        lb1 = Label(root,text="Pregnancies    :",bg="white")
        lb1.config(font=("Times New Roman",12))
        lb1.place(x=20,y=190)
        text1 = tk.Entry(root,width=10,relief='solid',borderwidth=1,bg='snow')
        text1.config(state=DISABLED)
        text1.place(x=185,y=193)
        #2
        lb1 = Label(root,text="Glucose        :",bg="white")
        lb1.config(font=("Times New Roman",12))
        lb1.place(x=20,y=220)
        lb1 = Label(root,text="70-180 mg/dl",bg="white")
        lb1.config(font=("Times New Roman",9))
        lb1.place(x=250,y=220)
        text2 = tk.Entry(root,width=10,relief='solid',borderwidth=1,bg='snow')
        text2.place(x=185,y=223)
        #3
        lb1 = Label(root,text="Blood Pressure :",bg="white")
        lb1.config(font=("Times New Roman",12))
        lb1.place(x=20,y=250)
        lb1 = Label(root,text="89-140mm HG",bg="white")
        lb1.config(font=("Times New Roman",9))
        lb1.place(x=250,y=250)
        text3 = tk.Entry(root,width=10,relief='solid',borderwidth=1,bg='snow')
        text3.place(x=185,y=253)
        #4
        lb1 = Label(root,text="Skin Thickness :",bg="white")
        lb1.config(font=("Times New Roman",12))
        lb1.place(x=20,y=280)
        lb1 = Label(root,text="10-50mm",bg="white")
        lb1.config(font=("Times New Roman",9))
        lb1.place(x=250,y=280)
        text4 = tk.Entry(root,width=10,relief='solid',borderwidth=1,bg='snow')
        text4.place(x=185,y=283)
        #5
        lb1 = Label(root,text="Insulin           :",bg="white")
        lb1.config(font=("Times New Roman",12))
        lb1.place(x=340,y=190)
        lb1 = Label(root,text="15-276mu U/mL",bg="white")
        lb1.config(font=("Times New Roman",9))
        lb1.place(x=570,y=190)
        text5 = tk.Entry(root,width=10,relief='solid',borderwidth=1,bg='snow')
        text5.place(x=505,y=193)
        #6
        lb1 = Label(root,text="BMI               :",bg="white")
        lb1.config(font=("Times New Roman",12))
        lb1.place(x=340,y=220)
        lb1 = Label(root,text="10-50",bg="white")
        lb1.config(font=("Times New Roman",9))
        lb1.place(x=570,y=220)
        text6 = tk.Entry(root,width=10,relief='solid',borderwidth=1,bg='snow')
        text6.place(x=505,y=223)
        #7
        lb1 = Label(root,text="Pedigree Function :",bg="white")
        lb1.config(font=("Times New Roman",12))
        lb1.place(x=340,y=250)
        text7 = tk.Entry(root,width=10,relief='solid',borderwidth=1,bg='snow')
        text7.place(x=505,y=253)
        #8
        lb1 = Label(root,text="Age               :",bg="white")
        lb1.config(font=("Times New Roman",12))
        lb1.place(x=340,y=280)
        lb1 = Label(root,text="30-80",bg="white")
        lb1.config(font=("Times New Roman",9))
        lb1.place(x=570,y=280)
        text8 = tk.Entry(root,width=10,relief='solid',borderwidth=1,bg='snow')
        text8.place(x=505,y=283)

        lb1 = Label(root,text="Name User:",bg="white")
        lb1.config(font=("Times New Roman",12))
        lb1.place(x=100,y=340)
        text9 = tk.Entry(root,width=30,relief='solid',borderwidth=1,bg='snow')
        text9.place(x=200,y=340)
        #check button
        chkbtn = Button(root,text="Check Result",relief=FLAT,bg='orange',width=14,command=check,cursor='hand2')
        chkbtn.config(font=("Times New Roman",10))
        chkbtn.place(x=485,y=120)
        
    #choose gender
    def gen():
        lb2 = Label(root,text="Select Gender",bg="white")
        lb2.config(font=("Times New Roman",14))
        lb2.place(x=20,y=140)
        choice=Combobox(root)
        choice['values']=("-----Select-----","Male","Female")
        choice.place(x=185,y=140)
        choice.current(0)
        def cho():
            if choice.get()=="Male":
                Mdiab()
            elif choice.get()=="Female":
                Fdiab()
        Button(root,text="Next",command=cho,relief=FLAT,bg='orange',width=10,cursor='hand2').place(x=350,y=140)
        
    #diab()
    choice=Combobox(root)
    choice['values']=("-----Select-----","Diabetes")
    choice.place(x=185,y=110)
    choice.current(0)
    def cho():
        if choice.get()=="Diabetes":
            gen()
    Button(root,text="Next",command=cho,relief=FLAT,bg='orange',width=10,cursor='hand2').place(x=350,y=110)
    
    #footer
    foot = PhotoImage(file="images3.png")
    footlb = Label(root,image= foot,text="",bg='white')
    footlb.place(x=380,y=220)

    root.mainloop()





welcome()
